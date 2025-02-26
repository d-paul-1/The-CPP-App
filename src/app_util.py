# This file contains Utility Functions that are used all over the app

import os
import sys
from datetime import datetime
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
import pandas as pd

COLOR_PALETTE = {
    "maroon": "C00000",
    "yellow":"FFC000",

    #Banner Colors
    "lavender":"D5B8EA",
    "pink": "FF8FFF",
    "blue": "BDD7F1",
    "peach": "F9D2BD",
    "green": "D6EDBD",
    "cyan": "D9FFFF",
    
    }

def resource_path(relative_path):
    """ Get the absolute path to a resource, works for dev and for PyInstaller """
    base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base_path, relative_path)


def get_save_name(spreadsheet_name):
    """ This Function forms the new file name using the nomencalture which is {filename_currentDate_cuurentTime)

    Args:
        spreadsheet_name (String): Spreadsheet Base name

    Returns:
        String: Formatted name of file path to be saved
    """
    current_datetime = datetime.now().strftime("%b-%d-%Y_%H-%M-%S")
    return f"{spreadsheet_name}_{current_datetime}.xlsx"


def apply_color(cell, color_name):
    """Applies a predefined color to a cell."""
    color_code = COLOR_PALETTE.get(color_name, "FFFFFF")  # Default to white if not found
    fill = PatternFill(start_color=color_code, end_color=color_code, fill_type="solid")
    cell.fill = fill

def read_excel_table(excel_file, sheet_name, table_range):
    """ This function is used to read the values inside any specified range

    Args:
        excel_file (_type_): Excel spreadsheet path
        sheet_name (_type_): Tab name
        table_range (_type_): Range of the table(can be found using extract ramge function)

    Returns:
        _type_: _description_
    """
    # Load Excel workbook
    wb = load_workbook(excel_file, read_only=True)
    try:
        # Select the worksheet
        ws = wb[sheet_name]

        # Read data from the specified table range into a list of lists
        table_data = []
        for row in ws[table_range]:
            table_data.append([cell.value for cell in row])

        # Convert the list of lists to a DataFrame
        df = pd.DataFrame(table_data[1:], columns=table_data[0])
    finally:
        # Ensure the workbook is closed
        wb.close()

    return df


def extract_table_boundaries(file_path):
    """ Returns table boundaries from a given sheet

    Args:
        file_path (_type_):Path of the Spreadsheet

    Returns:
        _type_: A List of Table Boundaries
    """
    # Load the workbook
    wb = load_workbook(file_path, data_only=True)

    # List of Table Boundaries
    table_ranges= []

    # Loop through all worksheets
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]

        # Check for tables in the sheet
        if hasattr(sheet, 'tables') and sheet.tables:
            for table_name in sheet.tables.keys():
                table = sheet.tables[table_name]  # Retrieve the table object
                if hasattr(table, 'ref'):  # Safely access ref attribute
                    table_range = table.ref
                    table_ranges.append(table_range)

    return table_ranges