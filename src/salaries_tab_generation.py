# This file contains all the required functions for the stlying, formatting and generation of the salraies tab in the Payroll Spreadsheet

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, NamedStyle, Border, Side
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime 
import os


def generate_salaries_tab(master_data_input_sheet_path, save_path, pay_periods):

    wb = Workbook()
    ws = wb.active
    ws.title = "SALARIES"
    ws.sheet_properties.tabColor = "C00000"
    add_presets(ws)
    final_save_path = os.path.join(save_path,get_save_name("Payroll_Spreadsheet"))
    wb.save(final_save_path)
    return final_save_path



def add_presets(ws):
    """ This function adds all the the presets that need to be hardcoded
    """


def get_save_name(spreadsheet_name):
    """ This Function forms the new file name using the nomencalture which is {filename_currentDate_cuurentTime)

    Args:
        spreadsheet_name (_type_): Spreadsheet Base name

    Returns:
        _type_: Name of file path to be saved
    """
    current_datetime = datetime.now().strftime("%b-%d-%Y_%H-%M-%S")
    return f"{spreadsheet_name}_{current_datetime}.xlsx"