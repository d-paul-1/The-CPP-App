

# This class is responsible for generating the payroll spreadsheet with both the salaries tab and the staff tab.



import os

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, NamedStyle, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

from datetime import datetime
import pandas as pd

COLOR_PALETTE = {
    "maroon": "C00000",

    "yellow":"FFC000",

    #Banner Colors
    "lavender":"D5B8EA",
    "pink": "FF8FFF",
    "blue": "BDD7F1",
    "peach": "F9D2BD",
    "green": "D6EDBD",
    "cyan": "D9FFFF",
    
    }


# FORMATING STYLES

# General format
general_style = NamedStyle(name="general")
general_style.number_format = 'General'

# Accounting format
accounting_style = NamedStyle(name="accounting")
accounting_style.number_format = '_("$"* #,##0.00_);_("$"* \\(#,##0.00\\);_("$"* "-"??_);_(@_)'


percentage_style = NamedStyle(name="percentage")
percentage_style.number_format = '0.00%'

date_style = NamedStyle(name="date_style")
date_style.number_format = 'mm/dd/yyyy'

# accounting formating
accounting_style = '_("$"* #,##0.00_);_("$"* \\(#,##0.00\\);_("$"* "-"??_);_(@_)'

general_style = NamedStyle(name="general")
general_style.number_format = 'General'

HEADER_ROWS = []  # Global list to store header row indices



def generate_payroll_spreadsheet(save_directory, inputsheet_two_path, pay_schedule,num_pay_periods):

    print(inputsheet_two_path)


    # Convert Pay Period Dates to datetime for processing
    pay_schedule['Start Date'] = pd.to_datetime(pay_schedule['Pay Period Dates'].str.split(' - ').str[0], format='%m/%d/%Y')
    pay_schedule['End Date'] = pd.to_datetime(pay_schedule['Pay Period Dates'].str.split(' - ').str[1], format='%m/%d/%Y')

    # Extracting Y1, Y2, Start, and End
    y1 = pay_schedule['Start Date'].min().year
    y2 = y1 + 1
    start_date = pay_schedule['Start Date'].min().strftime('%m/%d/%Y')
    end_date = pay_schedule['End Date'].max().strftime('%m/%d/%Y')

    # Displaying the values
    print(f"Y1 = {y1}")
    print(f"Y2 = {y2}")
    print(f"Start = {start_date}")
    print(f"End = {end_date}")


    data = {
        "Parameters": ["number of pay periods", "year 1 numeric" , "year 1 string", "year 2 numeric" , "year 2 string","academic_fringe",
                    "university_fringe","student_fringe","year_start","year_end"],
        "Values": [f"{num_pay_periods}" , f"{y1 % 100}", f"FY {y1 % 100}",  f"{y2 % 100}", f"FY {y2 % 100}","N/A Too be filled",
                "N/A Too be filled","N/A Too be filled",f"{start_date}",f"{end_date}"]
    }
    params_df = pd.DataFrame(data)

    final_save_path = os.path.join(save_directory,get_save_name(f"{params_df.loc[params_df['Parameters'] == 'year 2 string', 'Values'].iloc[0]} Payroll Spreadsheet "))

    wb= Workbook()
    generate_salaries_tab(wb,inputsheet_two_path,params_df)
    generate_staff_tab(wb,inputsheet_two_path, params_df)

    wb.save(final_save_path)
    wb.close
    
    return final_save_path


def generate_salaries_tab(wb,input_sheet_two_path, params_df):

    ws1 = wb.active
    ws1.title = f"{params_df.loc[params_df['Parameters'] == 'year 2 string', 'Values'].iloc[0]} SALARIES"
    ws1.sheet_properties.tabColor = "C00000"

    salaries_tab_presets(ws1,wb,input_sheet_two_path, params_df)




def salaries_tab_presets(ws,wb,input_sheet_two_path, params_df):

    # Set width of the first 200 columns and height of the first 200 rows
    for col in range(1, 201):  # Columns A to GR
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
    for row in range(3, 201):  # Rows 1 to 200
        ws.row_dimensions[row].height = 25

    ws.column_dimensions[openpyxl.utils.get_column_letter(1)].width = 26
    ws.column_dimensions[openpyxl.utils.get_column_letter(3)].width = 7

    # TITLE SETTINGS
    ws.row_dimensions[3].height= 28.50
    title_cell = ws.cell(row=3, column=1)
    title_cell.value = f"{params_df.loc[params_df['Parameters'] == 'year 2 string', 'Values'].iloc[0]} SALARY & FRINGE W/COLAS"
    title_cell.font = Font(size=22, bold=True)

    # TIME STAMP SETTINGS
    time_stamp_cell = ws.cell(row=3, column=5)
    apply_color(time_stamp_cell, "maroon")
    time_stamp_cell.value = "Generated on " + datetime.now().strftime("%b-%d-%Y") + " at " + datetime.now().strftime("%H:%M:%S")
    time_stamp_cell.font = Font(size=14, color="FFFFFF", bold=True)  
    ws.merge_cells(start_row=3, start_column=5, end_row=3, end_column=7)

    # FY BANNER SETTINGS
    ws.row_dimensions[1].height = 50.25
    banner =""
    for i in range(1,100):
        if i % 2 == 0:
            banner = banner + "    -    "
        else:
            banner = banner + f"{params_df.loc[params_df['Parameters'] == 'year 2 string', 'Values'].iloc[0]}"

    banner_cell = ws.cell(row=1, column=1)
    banner_cell.value = banner
    banner_cell.font=Font(name = "Ariel", size=18, color= "2F75B5" , bold=  True)
    banner_cell.alignment = Alignment(vertical="center")

    # banner Color 
    for col in range(1, 25):
        cell = ws.cell(row=1, column=col)
        apply_color(cell,get_banner_color(params_df.loc[params_df['Parameters'] == 'year 2 string', 'Values'].iloc[0]))

    # PRESETS FOR LINES 5-9

    font_blue59 = Font(color="0000CC",bold=  True)

    fringe_df = retrieve_table(input_sheet_two_path,"USER INPUT",["fringe"])
    fringe_df = fringe_df.dropna(how="all")
    

    if not fringe_df.empty:
        fringe_df = fringe_df.iloc[[-1]]
    else:
        print("The fringe_df DataFrame is empty!")

    

    
 

    cell_presets = {
    5: {
        1: {"value": f"{params_df.loc[params_df['Parameters'] == 'year 2 string', 'Values'].iloc[0]} Fringe - Academic Staff", "font": font_blue59},                                                               # Column A (1)
        2: {"value": fringe_df.iloc[0,2], "style": percentage_style, "fill": "cyan"},     # Column C (3)
        4: {"value": f"{params_df.loc[params_df['Parameters'] == 'year 2 string', 'Values'].iloc[0]} Fiscal Year Start", "font": font_blue59},                                                                     # Column F (6)
        5: {"value": params_df.loc[params_df["Parameters"] == "year_start", "Values"].iloc[0], "style": date_style, "fill": "cyan"},                # Column G (7)
    },
    6: {
        1: {"value": f"{params_df.loc[params_df['Parameters'] == 'year 2 string', 'Values'].iloc[0]} Fringe - Grad Students", "font": font_blue59},                                                             # Column A (1)
        2: {"value": fringe_df.iloc[0,3], "style": percentage_style, "fill": "cyan"},   # Column C (3)
        4: {"value": f"{params_df.loc[params_df['Parameters'] == 'year 2 string', 'Values'].iloc[0]} Fiscal Year End", "font": font_blue59},                                                                       # Column F (6)
        5: {"value": params_df.loc[params_df["Parameters"] == "year_end", "Values"].iloc[0], "style": date_style, "fill": "cyan"},                  # Column G (7)
    },
    7: {
        1: {"value": f"{params_df.loc[params_df['Parameters'] == 'year 2 string', 'Values'].iloc[0]} Fringe - Student hourly", "font": font_blue59},                                                                # Column A (1)
        2: {"value": fringe_df.iloc[0,4], "style": percentage_style, "fill": "cyan"},      # Column C (3)
        4: {"value": f"{params_df.loc[params_df['Parameters'] == 'year 2 string', 'Values'].iloc[0]} # Total Days", "font": font_blue59},                                                                          # Column F (6)
        5: {"value": "=(E6-E5)+1", "style": general_style,"fill": "cyan"},                                                                                                 # Column G (7)
    },
    9: {
        4: {"value": f"{params_df.loc[params_df['Parameters'] == 'year 2 string', 'Values'].iloc[0]} PAY PERIODS", "font": font_blue59},                                                                           # Column F (6)
        5: {"value": params_df.loc[params_df["Parameters"] == "number of pay periods", "Values"].iloc[0], "fill": "cyan"},                          # Column G (7)
    },
    }

    # Iterate over the dictionary to apply settings
    for row, columns in cell_presets.items():
        for col, settings in columns.items():
            cell = ws.cell(row=row, column=col)  # Use numeric column directly
            if "value" in settings:
                cell.value = settings["value"]
            if "font" in settings:
                cell.font = settings["font"]
            if "style" in settings:
                cell.style = settings["style"]
            if "fill" in settings:
                apply_color(cell, settings["fill"])

    print_functions = [
      print_salaried_employee,
      print_lump_sum_employee,
      print_undergrad_employee,
      print_grad_employee,
      print_other_dep_employee
    ]

    current_row = 12  # starting row
    for print_func in print_functions:
            current_row = print_func(ws,wb,input_sheet_two_path,params_df,fringe_df, current_row)


      # Format specific columns
    date_format_columns = ['E', 'F']  # Add columns that need date formatting
    accounting_format_columns = ['G','H', 'I', 'K','L', 'M', 'N', 'O', 'P', 'Q']  # Add columns that need accounting formatting

    # Apply date formatting
    for col in date_format_columns:
        format_column(ws, col, date_style)

    # Apply accounting formatting
    for col in accounting_format_columns:
        format_column(ws, col, accounting_style)

    ws["E7"].number_format = general_style.number_format


    # Add this line at the end
    format_headers(ws)
    

def print_salaried_employee(ws,wb,input_sheet_two_path,params_df,fringe_df, start_row):
    # Printing the Heading 
    print_header(ws,"SALARIED EMPLOYEES", start_row)
    start_row_offset = start_row +2 # this the so that the table doesnt print on the same line as the header

    df = retrieve_table(input_sheet_two_path,"USER INPUT",["employee_data_Salaried_Employees"])
    #df = df.dropna(how="all")  uncomment if you wanr to remove the spacing
    new_df = restructure_dataframe(df,params_df, input_sheet_two_path)
    
    # Write the table to the workbook
    write_tables_to_excel(
        ws=ws,
        sheet_name="USER INPUT",
        tables_list=[new_df],
        start_row=start_row_offset,
        base_table_name="salaried_employee",
        buffer=3
    )

    # Return the row index after the table and buffer
    return start_row_offset + len(new_df) + 3  # 3 is the buffer

def print_lump_sum_employee(ws,wb,input_sheet_two_path,params_df,fringe_df, start_row):
    # Printing the Heading 
    print_header(ws,"LUMP SUM/NO FRINGE EMPLOYEES - INSTRUCTORS", start_row)
    start_row_offset = start_row +2 # this the so that the table doesnt print on the same line as the header

    df = retrieve_table(input_sheet_two_path,"USER INPUT",["employee_data_Lump_Sum_No_Fringe_Employees___Instructors"])
    #df = df.dropna(how="all")  uncomment if you wanr to remove the spacing
    new_df = restructure_dataframe(df,params_df, input_sheet_two_path)
    
    # Write the table to the workbook
    write_tables_to_excel(
        ws=ws,
        sheet_name="USER INPUT",
        tables_list=[new_df],
        start_row=start_row_offset,
        base_table_name="lump_sum_employee",
        buffer=3
    )

    # Return the row index after the table and buffer
    return start_row_offset + len(df) + 3  # 3 is the buffer


def print_undergrad_employee(ws,wb,input_sheet_two_path,params_df,fringe_df,start_row):
    # Printing the Heading 
    print_header(ws,"UNDERGRADUATE STUDENT EMPLOYEES", start_row)
    start_row_offset = start_row +2 # this the so that the table doesnt print on the same line as the header

    df = retrieve_table(input_sheet_two_path,"USER INPUT",["employee_data_Undergraduate_Student_Employees"])
    #df = df.dropna(how="all")  uncomment if you wanr to remove the spacing
    new_df = restructure_dataframe(df,params_df, input_sheet_two_path)
    
    # Write the table to the workbook
    write_tables_to_excel(
        ws=ws,
        sheet_name="USER INPUT",
        tables_list=[new_df],
        start_row=start_row_offset,
        base_table_name="Undergraduate_student_employee",
        buffer=3
    )

    # Return the row index after the table and buffer
    return start_row_offset + len(df) + 3  # 3 is the buffer

def print_grad_employee(ws,wb,input_sheet_two_path,params_df,fringe_df, start_row):
    # Printing the Heading 
    print_header(ws,"GRADUATE STUDENT EMPLOYEES", start_row)
    start_row_offset = start_row +2 # this the so that the table doesnt print on the same line as the header


    df = retrieve_table(input_sheet_two_path,"USER INPUT",["employee_data_Graduate_Student_Employees"])
    #df = df.dropna(how="all")  uncomment if you wanr to remove the spacing
    new_df = restructure_dataframe(df,params_df, input_sheet_two_path)
    
    # Write the table to the workbook
    write_tables_to_excel(
        ws=ws,
        sheet_name="USER INPUT",
        tables_list=[new_df],
        start_row=start_row_offset,
        base_table_name="Graduate_student_employee",
        buffer=3
    )


    # Return the row index after the table and buffer
    return start_row_offset + len(df) + 3  # 3 is the buffer


def print_other_dep_employee(ws,wb,input_sheet_two_path,params_df,fringe_df, start_row):
    # Printing the Heading 
    print_header(ws,"EMPLOYEES PAID BY OTHER DEPARTMENTS", start_row)
    start_row_offset = start_row +2 # this the so that the table doesnt print on the same line as the header

    df = retrieve_table(input_sheet_two_path,"USER INPUT",["employee_data_Employees_Paid_By_Other_Departments"])
    #df = df.dropna(how="all")  uncomment if you wanr to remove the spacing
    new_df = restructure_dataframe(df,params_df, input_sheet_two_path)
    
    # Write the table to the workbook
    write_tables_to_excel(
        ws=ws,
        sheet_name="USER INPUT",
        tables_list=[new_df],
        start_row=start_row_offset,
        base_table_name="other_departments_employee",
        buffer=3
    )

    # Return the row index after the table and buffer
    return start_row_offset + len(df) + 3  # 3 is the buffer


def print_header(ws, heading, row):
    """
    Writes a single formatted heading to the specified row in the workbook.

    Parameters:
        wb (Workbook): The openpyxl Workbook object.
        heading (str): The text to display as the heading.
        row (int): The row number where the heading should be placed.
    """

    # Add the row number to our global list
    HEADER_ROWS.append(row+2) # +2 because the table is 2 rows below the header.

    # Define style
    heading_font = Font(bold=True, size=12)
    heading_alignment = Alignment(horizontal="left", vertical="center")

    # Write the heading to column 1 of the specified row
    cell = ws.cell(row=row, column=1, value=heading)
    cell.font = heading_font
    cell.alignment = heading_alignment

    # Color the full row (up to column 30)
    for col_num in range(1, 25):
        apply_color(ws.cell(row=row, column=col_num), "yellow")





def generate_staff_tab(wb,input_sheet_two_path,params_df):

    ws2 = wb.create_sheet(title=f"{params_df.loc[params_df['Parameters'] == 'year 2 string', 'Values'].iloc[0]} STAFF")
    ws2.sheet_properties.tabColor = "0000CC"


def get_save_name(spreadsheet_name):

    current_datetime = datetime.now().strftime("%b-%d-%Y_%H-%M-%S")
    return f"{spreadsheet_name}_{current_datetime}.xlsx"


def get_banner_color(curr_year):
    """This functions inputs the current year and returns the color required for the banner

    Args:
        curr_year (String): current year

    Returns:
        String: respective color for year
    """
    end=int(curr_year[-1])
    if end == 4 or end == 9:
        return "green"

    elif end == 5 or end == 0:
        return "pink"

    elif end == 6  or end == 1:
        return "blue"

    elif end == 7  or end == 2 :
        return "peach"

    elif end == 8 or end == 3 :
        return "lavender"
    
def apply_color(cell, color_name):
    """Applies a predefined color to a cell."""
    color_code = COLOR_PALETTE.get(color_name, "FFFFFF")  # Default to white if not found
    fill = PatternFill(start_color=color_code, end_color=color_code, fill_type="solid")
    cell.fill = fill

def retrieve_table(file_path, sheet_name, table_names):
    from openpyxl import load_workbook
    import pandas as pd

    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb[sheet_name]

        dataframes = {}
        for table_name in table_names:
            table = ws.tables.get(table_name)
            if table:
                if isinstance(table, str):
                    # Table is stored as a string
                    print(f"Warning: Table '{table_name}' appears as a string. Check openpyxl version.")
                    continue

                # Get table range and convert to DataFrame
                start_cell, end_cell = table.ref.split(":")
                data = ws[start_cell:end_cell]
                headers = [cell.value for cell in data[0]]
                rows = [[cell.value for cell in row] for row in data[1:]]
                dataframes[table_name] = pd.DataFrame(rows, columns=headers)
            else:
                print(f"Table '{table_name}' not found in sheet '{sheet_name}'.")

        # If only one table name is provided, return its DataFrame directly
        if len(table_names) == 1:
            return dataframes.get(table_names[0], pd.DataFrame())
        return dataframes

    except Exception as e:
        print(f"An error occurred: {e}")
        return None

def extract_table_boundaries(file_path):
    """ Returns table boundaries from a given sheet

    Args:
        file_path (_type_):Path of the Spreadsheet

    Returns:
        _type_: A List of Table Boundaries
    """
    # Load the workbook
    wb = load_workbook(file_path, data_only=True)

    # List of Table Boundaries
    table_ranges= []

    # Loop through all worksheets
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]

        # Check for tables in the sheet
        if hasattr(sheet, 'tables') and sheet.tables:
            for table_name in sheet.tables.keys():
                table = sheet.tables[table_name]  # Retrieve the table object
                if hasattr(table, 'ref'):  # Safely access ref attribute
                    table_range = table.ref
                    table_ranges.append(table_range)

    return table_ranges


def read_excel_table(excel_file, sheet_name, table_range):
    """ This function is used to read the values inside any specified range

    Args:
        excel_file (_type_): Excel spreadsheet path
        sheet_name (_type_): Tab name
        table_range (_type_): Range of the table(can be found using extract ramge function)

    Returns:
        _type_: _description_
    """
    # Load Excel workbook
    wb = load_workbook(excel_file, read_only=True)
    try:
        # Select the worksheet
        ws = wb[sheet_name]

        # Read data from the specified table range into a list of lists
        table_data = []
        for row in ws[table_range]:
            table_data.append([cell.value for cell in row])

        # Convert the list of lists to a DataFrame
        df = pd.DataFrame(table_data[1:], columns=table_data[0])
    finally:
        # Ensure the workbook is closed
        wb.close()

    return df


def get_colas(data_ips2):
    colas_info_table = retrieve_table(data_ips2,"USER INPUT",["colas_info"])
    colas_info_table = colas_info_table.dropna(how="all")
    colas_info_table= colas_info_table.iloc[[-1]]
    return colas_info_table.iloc[0,1]

def write_tables_to_excel(ws, sheet_name, tables_list, start_row, base_table_name, mode=1, start_col= 1, buffer=3):
    """
    Write a list of DataFrames to an Excel sheet as tables.

    Args:
        wb (Workbook): Excel workbook object.
        sheet_name (str): Name of the sheet to write to.
        tables_list (list): List of DataFrames to write.
        start_row (int): Starting row for the first table.
        buffer (int): Number of blank rows between tables.
    """

    
    current_row = start_row

    for idx, df in enumerate(tables_list):
        # Ensure DataFrame is not empty
        if df.empty:
            print(f"Skipping empty DataFrame at index {idx}")
            continue

        # Ensure DataFrame has unique column headers
        if len(df.columns) != len(set(df.columns)):
            print(f"DataFrame at index {idx} has duplicate column headers. Skipping.")
            continue

        if mode==2:
            # Use the first column header as the base for the table name
            first_header = str(df.columns[0])
            # Sanitize the table name
            sanitized_header = ''.join(e if e.isalnum() else '_' for e in first_header)
            table_name = f"{base_table_name}_{sanitized_header}"
        else :
            table_name = f"{base_table_name}"

        # Define the range for the table
        start_col = start_col  
        end_col = start_col + len(df.columns) - 1
        end_row = current_row + len(df)

        table_range = f"{get_column_letter(start_col)}{current_row}:{get_column_letter(end_col)}{end_row}"

        # Write the DataFrame to the sheet
        for row in dataframe_to_rows(df, index=False, header=True):
            for col_idx, value in enumerate(row, start=start_col):
                ws.cell(row=current_row, column=col_idx, value=value)
            current_row += 1

        # Add buffer
        current_row += buffer

        # Validate the table range before creating the table
        if len(df) > 0 and len(df.columns) > 0:
            try:
                # Create an Excel table
                table = Table(displayName=table_name, ref=table_range)
                style = TableStyleInfo(
                    name="TableStyleMedium9",  # Choose any predefined Excel style
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False,
                )
                table.tableStyleInfo = style
                ws.add_table(table)
                print(f"Successfully added table: {table_name} with range: {table_range}")
            except ValueError as e:
                print(f"Error creating table {table_name}: {e}")
        else:
            print(f"Skipping DataFrame at index {idx} due to invalid dimensions.")

    print(f"All tables written to sheet '{sheet_name}' successfully.")


def restructure_dataframe(df,  params_df, data_ips2):
    
    colas_table = retrieve_table(data_ips2,"USER INPUT",["colas_eligible"])
    colas_table = colas_table.dropna(how="all")
    colas_table = colas_table.T.reset_index()
    colas_table.columns = ['Name', 'Eligible']

    # Apply case-sensitive mapping
    colas_table['Eligible'] = colas_table['Eligible'].map({'Y': True, 'N': False, 'y': True, 'n': False})

    # List of new columns to be added
    new_columns = [
        f"{df.columns[0]}",
        "Employee Number",
        "List Order",
        "Name",
        "Start Date",
        "End Date",
        f"{params_df.loc[params_df['Parameters'] == 'year 1 string', 'Values'].iloc[0]} END BASE Pay",
        "2% COLA / Raise / FR %",
        f"{params_df.loc[params_df['Parameters'] == 'year 2 string', 'Values'].iloc[0]} START BASE Pay",
        "% FTE",
        f"{params_df.loc[params_df['Parameters'] == 'year 2 string', 'Values'].iloc[0]} Pay by % Appt",
        "FRINGE",
        f"{params_df.loc[params_df['Parameters'] == 'year 2 string', 'Values'].iloc[0]} TOTAL Pay by % Appt",
        "Sub-Total Actual Pay",
        "Bonus / Lump Sum Pay",
        f"{params_df.loc[params_df['Parameters'] == 'year 2 string', 'Values'].iloc[0]} ACTUAL PAY",
        f"{params_df.loc[params_df['Parameters'] == 'year 2 string', 'Values'].iloc[0]} END BASE Pay = FY{str(int(params_df.loc[params_df['Parameters'] == 'year 2 numeric', 'Values'].iloc[0]) + 1)} START BASE Pay"
        
    ]
    
    new_df = pd.DataFrame(columns=new_columns)

    new_df[new_df.columns[1]] = df[df.columns[1]]
    new_df['List Order'] = df[df.columns[2]]
    new_df['Name'] = df[df.columns[3]]
    new_df['Start Date'] = df[df.columns[8]]
    new_df['End Date'] = df[df.columns[9]]
    new_df[new_df.columns[6]] = df[df.columns[6]]

    # Loop through the names and check if they are eligible in colas_table
    for idx, row in new_df.iterrows():
        name = row['Name']
        if name in colas_table['Name'].values:  # Check if the name exists in colas_table
            eligible = colas_table.loc[colas_table['Name'] == name, 'Eligible'].iloc[0]
            if eligible:  # If the person is eligible for COLA
                # Apply the 2% COLA to the FY24 END BASE Pay column (or the appropriate column)
                # Use loc to directly assign values
                new_df.loc[idx, '2% COLA / Raise / FR %'] = f"=INDIRECT(ADDRESS(ROW(), COLUMN()-1))*{1+ get_colas(data_ips2)}" # 2% COLA



    
    new_df[new_df.columns[8]] = f"=IF(ISBLANK(INDIRECT(ADDRESS(ROW(), COLUMN()-1))),INDIRECT(ADDRESS(ROW(), COLUMN()-2)),INDIRECT(ADDRESS(ROW(), COLUMN()-1)))"  
    new_df['% FTE'] = df[df.columns[4]]
    new_df[new_df.columns[10]] = f"=INDIRECT(ADDRESS(ROW(), COLUMN()-1))*INDIRECT(ADDRESS(ROW(), COLUMN()-2))"

    # Populate 'FRINGE' based on the keywords in the 'Fringe Type' column from df
    for idx, row in new_df.iterrows():
        # Get the fringe type from the original 'df'
        fringe_type = str(df.loc[idx, 'Fringe Type']).lower()  # Convert to lower case for case-insensitive comparison
        
        # Check for the keywords and set the correct reference
        if 'academic' in fringe_type:
            new_df.loc[idx, 'FRINGE'] = f"=$B$5 * INDIRECT(ADDRESS(ROW(), COLUMN()-1))"  # Academic Staff
        elif 'grad' in fringe_type:
            new_df.loc[idx, 'FRINGE'] = f"=$B$6 * INDIRECT(ADDRESS(ROW(), COLUMN()-1))"  # Grad Students
        elif 'hourly' in fringe_type:
            new_df.loc[idx, 'FRINGE'] = f"=$B$7 * INDIRECT(ADDRESS(ROW(), COLUMN()-1))"  # Student Hourlies
        else:
            new_df.loc[idx, 'FRINGE'] = f"N/A"  # Optional: If no match is found, set to 'N/A' or any default value




    new_df[new_df.columns[12]] = f"=SUM(INDIRECT(ADDRESS(ROW(), COLUMN()-1)):INDIRECT(ADDRESS(ROW(), COLUMN()-2)))"
    new_df['Sub-Total Actual Pay'] = f"=INDIRECT(ADDRESS(ROW(), COLUMN()-1))*((INDIRECT(ADDRESS(ROW(), COLUMN()-8))-INDIRECT(ADDRESS(ROW(), COLUMN()-9))+1)/$E$7)"
    # new_df['Bonus / Lump Sum Pay'] =
    counter = 0

    # Loop through the rows of the DataFrame
    for idx in range(len(new_df)):
        counter += 1  # Increment the counter for each row
        
        # Apply the formulas every 4th row (counter % 4 == 1 for first, 5th, 9th, etc.)
        if counter % 4 == 1:
            # Apply the SUM formula for columns 15 and 16 on this row
            new_df.loc[idx, new_df.columns[15]] = f"=SUM(INDIRECT(ADDRESS(ROW(), COLUMN()-2)):INDIRECT(ADDRESS(ROW()+3, COLUMN()-1)))"
            new_df.loc[idx, new_df.columns[16]] = f"=LOOKUP(2,1/(INDIRECT(ADDRESS(ROW(), COLUMN()-8)):INDIRECT(ADDRESS(ROW()+3, COLUMN()-8))<>0),INDIRECT(ADDRESS(ROW(), COLUMN()-8)):INDIRECT(ADDRESS(ROW()+3, COLUMN()-8)))"

        

    
    return new_df


def format_headers(ws):
    """
    Applies text wrapping to all header rows from column A to Q.
    """
    wrap_alignment = Alignment(wrap_text=True, vertical='center')
    
    for row in HEADER_ROWS:
        for col in range(1, 18):  # Columns A to Q (1 to 17)
            cell = ws.cell(row=row, column=col)
            cell.alignment = wrap_alignment

        ws.row_dimensions[row].height= 45


def format_column(ws, column, style):
    """
    Format an entire column's number format while preserving other formatting
    
    Args:
        ws: worksheet
        column: column letter (e.g., 'A') or number (1)
        style: NamedStyle object or number format string
    """
    if isinstance(column, str):
        col_idx = openpyxl.utils.column_index_from_string(column)
    else:
        col_idx = column

    for row in range(1, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        # Only apply the number format, preserving other styles
        if isinstance(style, NamedStyle):
            cell.number_format = style.number_format
        else:
            cell.number_format = style