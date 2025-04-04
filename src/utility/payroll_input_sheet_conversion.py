# This file handles the code responsible for converting INPUT SHEET 1 TO 2

from tkinter import filedialog
import warnings
from pandas.errors import SettingWithCopyWarning  # Import the warning class

# Suppress the specific pandas warning
warnings.simplefilter(action='ignore', category=SettingWithCopyWarning)


from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Color
from datetime import datetime


import pandas as pd
import numpy as np
import re
import openpyxl


def generate_input_sheet_two(save_directory, input_sheet_one_path):
    """
    Generates 'Input Data Sheet 2' Excel workbook from a given input file and saves it to the specified directory.

    If no save directory is provided, prompts the user to select one. The function creates a new Excel workbook,
    populates it with presets and data extracted from the provided 'Input Sheet 1' file, and saves it with a
    timestamped filename.

    Args:
        save_directory (str): Path to the directory where the Excel file should be saved. If empty, prompts the user.
        input_sheet_one_path (str): File path to the existing 'Input Sheet 1' Excel file to pull data from.

    Returns:
        str: Full file path where the new Excel workbook was saved.
    """


    # Prompting the user to select save folder if not already selected
    if not save_directory:
        save_directory = filedialog.askdirectory(title="Select Folder to Save Payroll Spreadsheet") or ""
    
    # Creating Workbook
    wb = Workbook()

    # Formatting the Final Save Path
    final_save_path = os.path.join(save_directory,get_save_name(f"CPP Input Data Sheet 2 - ALL INPUT DATA"))

    # Populating the Excel Sheet
    create_presets(wb)

    # Calling Individual section funtions

    print_functions = [
        print_updates,
        print_fringe_table,
        print_term_leave_table,
        print_cpp_employees_and_salary_data,
        print_colas,
        print_funding_strings_and_accounting_salary,
        print_funding_strings_and_accounting_non_salary
    ]

    current_row = 3  # starting row
    for print_func in print_functions:
        if print_func == print_cpp_employees_and_salary_data:
            current_row = print_cpp_employees_and_salary_data(wb, current_row, input_sheet_one_path)
        elif print_func == print_funding_strings_and_accounting_salary:
           current_row = print_funding_strings_and_accounting_salary(wb, current_row, input_sheet_one_path)
        else:
            current_row = print_func(wb, current_row)

        
    # Saving and Closing the Excel Sheet
    wb.save(final_save_path)
    wb.close()

    # Returning the Final Save Path of Data Input Sheet 2
    return final_save_path
    

def create_presets(wb):
    # setting tab title
    ws = wb.active
    ws.title = f"USER INPUT"

    # TITLE SETTINGS
    ws.row_dimensions[1].height= 28.50
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"INPUT DATA TABLE 2 - ALL INPUT DATA"
    title_cell.font = Font(size=22, bold=True)

    # TIME STAMP SETTINGS
    time_stamp_cell = ws.cell(row=1, column=6)
    apply_color(time_stamp_cell, "maroon")
    time_stamp_cell.value = "Generated on " + datetime.now().strftime("%b-%d-%Y") + " at " + datetime.now().strftime("%H:%M:%S")
    time_stamp_cell.font = Font(size=14, color="FFFFFF", bold=True)  
    ws.merge_cells(start_row=1, start_column=6, end_row=1, end_column=8) 

    # Set width of the first 200 columns and height of the first 200 rows
    for col in range(1, 201):  # Columns A to GR
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 40
    for row in range(3, 201):  # Rows 1 to 200
        ws.row_dimensions[row].height = 25


def print_header(wb, heading, row):
    """
    Writes a single formatted heading to the specified row in the workbook.

    Parameters:
        wb (Workbook): The openpyxl Workbook object.
        heading (str): The text to display as the heading.
        row (int): The row number where the heading should be placed.
    """

    ws = wb.active  # Use the active worksheet

    # Define style
    heading_font = Font(bold=True, size=12)
    heading_alignment = Alignment(horizontal="left", vertical="center")

    # Write the heading to column 1 of the specified row
    cell = ws.cell(row=row, column=1, value=heading)
    cell.font = heading_font
    cell.alignment = heading_alignment

    # Color the full row (up to column 30)
    for col_num in range(1, 41):
        apply_color(ws.cell(row=row, column=col_num), "yellow")


def print_updates(wb, start_row):

    # Printing the Heading 
    print_header(wb,"UPDATES", start_row)
    start_row_offset = start_row +2 # this the so that the table doesnt print on the same line as the header

    # Create the DataFrame with 2 columns and 5 empty rows
    df = pd.DataFrame({
        "Date": [""] * 5,
        "Notes": [""] * 5
    })

    # Write the table to the workbook
    write_tables_to_excel(
        wb=wb,
        sheet_name="USER INPUT",
        tables_list=[df],
        start_row=start_row_offset,
        base_table_name="updates",
        buffer=3
    )

    # Return the row index after the table and buffer
    return start_row_offset + len(df) + 3  # 3 is the buffer

def print_fringe_table(wb, start_row):

    # Printing the Heading 
    print_header(wb,"FRINGE TABLE", start_row)
    start_row_offset = start_row +2 # this the so that the table doesnt print on the same line as the header

    # Create the DataFrame with 6 columns and 5 empty rows
    df = pd.DataFrame({
        "Fiscal Year": [""] * 5,
        "Effective Date": [""] * 5,
        "Academic Staff":[""] * 5,
        "Grad Students (PAs, RAs, TAs)":[""] * 5,
        "Student Hourlies":[""] * 5,
        "Post-Docs":[""] * 5
    })

    # Write the table to the workbook
    write_tables_to_excel(
        wb=wb,
        sheet_name="USER INPUT",
        tables_list=[df],
        start_row=start_row_offset,
        base_table_name="fringe",
        buffer=3
    )

    # Return the row index after the table and buffer
    return start_row_offset + len(df) + 3  # 3 is the buffer


def print_term_leave_table(wb, start_row):

    # Printing the Heading 
    print_header(wb,"TERM LEAVE TABLE", start_row)
    start_row_offset = start_row +2 # this the so that the table doesnt print on the same line as the header

    # Create the DataFrame with 2 columns and 5 empty rows
    df = pd.DataFrame({
        "Fiscal Year": [""] * 5,
        "Effective Date": [""] * 5,
        "Faculty & Academic Staff ":[""] * 5
    })

    # Write the table to the workbook
    write_tables_to_excel(
        wb=wb,
        sheet_name="USER INPUT",
        tables_list=[df],
        start_row=start_row_offset,
        base_table_name="term_leave",
        buffer=3
    )

    # Return the row index after the table and buffer
    return start_row_offset + len(df) + 3  # 3 is the buffer


def print_cpp_employees_and_salary_data(wb, start_row, input_sheet_1_path):

    # Printing the Heading 
    print_header(wb,"CPP EMPLOYEES & SALARY DATA", start_row)
    start_row_offset = start_row +2 # this the so that the table doesnt print on the same line as the header

    # Initialize an empty list to store DataFrames
    df_list = []

    # Extract table boundaries from the Excel file
    tables = extract_table_boundaries(input_sheet_1_path)

    # Loop through each table, process it, and append to the list
    for table in tables:
        # Read the table into a DataFrame
        df = read_excel_table(input_sheet_1_path, "USER INPUT", table)
        
        # Append the DataFrame to the list
        df_list.append(df)

    
    # This the the key matching dictionary for all the tables
    desired_columns_map = {
        "salaried employees": [
            "Employee Number", "List Order", "Name", "%FTE",
            "Actual Fringe Rate (104, 131, 136 accts)", "Salary",
            "Fringe Type", "Start Date (if new)", "End Date (if appl.)"
        ],
        "lump sum": [
            "Employee Number", "Teaching", "Name", "%FTE",
            "Home Dept", "Salary", "Fringe Type",
            "Start Date (if new)", "End Date (if appl.)"
        ],
        "undergraduate": [
            "Employee Number", " ", "Name", "%Appt",
            "  ", "Salary", "Fringe Type",
            "Start Date (if new)", "End Date (if appl.)"
        ],
        "graduate": [
            "Employee Number", " ", "Name", "%Appt",
            "  ", "Salary", "Fringe Type",
            "Start Date (if new)", "End Date (if appl.)"
        ],
        "other department": [
            "Employee Number", " ", "Name", "%Appt",
            "  ", "Salary", "Fringe Type",
            "Start Date (if new)", "End Date (if appl.)"
        ]
    }

    # Reverse the order of the list
    df_list = df_list[::-1]
    df_list = process_dfs_based_on_identifier(df_list,desired_columns_map)

    write_tables_to_excel(
        wb= wb,
        sheet_name="USER INPUT",
        tables_list=  df_list,
        start_row=start_row_offset,
        base_table_name="employee_data",
        buffer=3
    )

    # Return the row index after the table and buffer
    # Calculate total number of rows across all tables
    total_rows = sum(len(df) for df in df_list)

    # Total number of DataFrames/tables
    num_tables = len(df_list)

    # Return the final row index after all tables and buffer (3 rows per table)
    return start_row_offset + total_rows + (num_tables*3) + num_tables  # 3 is the buffer


def print_colas(wb, start_row):

    # Filter the DataFrame to get only salaried employees
    salaried_employees_df = all_employee_info[all_employee_info['Category'] == 'salaried employees']

    # Convert the names into columns with 2 empty rows
    salaried_employees_names = salaried_employees_df['Name'].values

    # Create a new DataFrame with 2 rows and the names as columns, using repetition for empty rows
    columns_dict = {name: [None] * 5 for name in salaried_employees_names}  # Repeat None for 2 rows

    # Create the new DataFrame
    salaried_employees_df = pd.DataFrame(columns_dict)


    # Printing the Heading 
    print_header(wb,"COLAS (Cost of Living Increases)", start_row)
    start_row_offset = start_row +3 # this the so that the table doesnt print on the same line as the header

    # Get the active sheet
    ws = wb['USER INPUT']

    # Add some text at start_row_offset, col 9, make it bold, and merge columns 9 to 12
    # Set the value first, before merging
    cell = ws.cell(row=start_row_offset - 1, column=4, value="EMPLOYEE ELIGIBLE?")

    # Merge cells from column 9 to 12 in the current row
    ws.merge_cells(start_row=start_row_offset - 1, start_column=4, end_row=start_row_offset - 1, end_column=3 + salaried_employees_df.shape[1])

    # Make the text bold
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")

    
    # Write the table to the workbook
    write_tables_to_excel(
        wb=wb,
        sheet_name="USER INPUT",
        tables_list=[salaried_employees_df],
        start_row=start_row_offset,
        base_table_name="colas",
        start_col=4,
        buffer=3
    )

    # Create the DataFrame with 2 columns and 5 empty rows
    df = pd.DataFrame({
        "Effective Date": [""] * 5,
        "Percentage ":[""] * 5
    })

    # Write the table to the workbook
    write_tables_to_excel(
        wb=wb,
        sheet_name="USER INPUT",
        tables_list=[df],
        start_row=start_row_offset,
        base_table_name="colas",
        buffer=3
    )

    # Return the row index after the table and buffer
    return start_row_offset + len(df) + 3  # 3 is the buffer


def print_funding_strings_and_accounting_salary(wb, start_row, input_sheet_1_path):


    # Printing the Heading 
    print_header(wb,"FUNDING STRINGS & ACCOUNTING -- SALARY", start_row)
    start_row_offset = start_row + 3 # this the so that the table doesnt print on the same line as the header

    # Get the active sheet
    ws = wb['USER INPUT']

    # Add some text at start_row_offset, col 9, make it bold, and merge columns 9 to 12
    # Set the value first, before merging
    cell = ws.cell(row=start_row_offset - 1, column=10, value="FUND ACCOUNTING")

    # Merge cells from column 9 to 12 in the current row
    ws.merge_cells(start_row=start_row_offset - 1, start_column=10, end_row=start_row_offset - 1, end_column=15)

    # Make the text bold
    cell.font = Font(bold=True, color="FF0000")
    cell.alignment = Alignment(horizontal="center", vertical="center")


    # Initialize an empty list to store DataFrames
    df_list = []

    # Extract table boundaries from the Excel file
    tables = extract_table_boundaries(input_sheet_1_path)

    # Loop through each table, process it, and append to the list
    for table in tables:
        # Read the table into a DataFrame
        df = read_excel_table(input_sheet_1_path, "USER INPUT", table)
        
        # Append the DataFrame to the list
        df_list.append(df)

    desired_columns_map = {
        "funding strings":["Department", "Color","Fund","Account","Sub-Account","Code","Description"]
    }

    # Reverse the order of the list
    df_list = df_list[::-1]
    df_list = process_dfs_based_on_identifier(df_list,desired_columns_map)

    write_tables_to_excel(
        wb= wb,
        sheet_name="USER INPUT",
        tables_list=  df_list,
        start_row=start_row_offset,
        base_table_name= "funding_strings",
        buffer=3
    )

    # Return the row index after the table and buffer
    # Calculate total number of rows across all tables
    total_rows = sum(len(df) for df in df_list)

    # Total number of DataFrames/tables
    num_tables = len(df_list)

    fund_acc_len = total_rows

    df_1 = pd.DataFrame({
        "EOY Balance": [""] * fund_acc_len,
        "Incoming Funds 1- Amount": [""] * fund_acc_len,
        "Incoming Funds 1- Source": [""] * fund_acc_len,
        "Incoming Funds 2- Amount": [""] * fund_acc_len,
        "Incoming Funds 3- Source": [""] * fund_acc_len,
        "Incoming Funds 3- Amount": [""] * fund_acc_len,
        "Incoming Funds 3- Source": [""] * fund_acc_len
    })

    write_tables_to_excel(
        wb=wb,
        sheet_name="USER INPUT",
        tables_list=[df_1],
        start_row=start_row_offset,
        base_table_name="funding_strings",
        start_col=10,
        buffer=3
    )

    employees_df = all_employee_info

    # Convert the names into columns with 2 empty rows
    employees_names = employees_df['Name'].values

    # Create a new DataFrame with 2 rows and the names as columns, using repetition for empty rows
    columns_dict = {name: [None] * fund_acc_len for name in employees_names}  # Repeat None for 2 rows

    # Create the new DataFrame
    employees_df = pd.DataFrame(columns_dict)

    cell_2 = ws.cell(row=start_row_offset - 1, column=17, value="%EFFORT PER EMPLOYEE PER FUND")

    # Merge cells from column 17 to (end of th employees df) in the current row
    ws.merge_cells(start_row=start_row_offset - 1, start_column=17, end_row=start_row_offset - 1, end_column=16 + employees_df.shape[1])

    # Make the text bold
    cell_2.font = Font(bold=True, color="FF0000")
    cell_2.alignment = Alignment(horizontal="center", vertical="center")


    write_tables_to_excel(
        wb= wb,
        sheet_name="USER INPUT",
        tables_list=  [employees_df],
        start_row=start_row_offset,
        base_table_name= "funding_strings",
        start_col= 17,
        buffer=3
    )



    # Return the final row index after all tables and buffer (3 rows per table)
    return start_row_offset + total_rows + (num_tables*3) + num_tables  # 3 is the buffer


def print_funding_strings_and_accounting_non_salary(wb, start_row):

    # Printing the Heading 
    print_header(wb, "FUNDING STRINGS & ACCOUNTING -- NON SALARY", start_row)
    start_row_offset = start_row + 3  # this ensures the table doesn't print on the same line as the header

    # Get the active sheet
    ws = wb['USER INPUT']

    # Add some text at start_row_offset, col 9, make it bold, and merge columns 9 to 12
    # Set the value first, before merging
    cell = ws.cell(row=start_row_offset - 1, column=9, value="FUND ACCOUNTING")

    # Merge cells from column 9 to 12 in the current row
    ws.merge_cells(start_row=start_row_offset - 1, start_column=9, end_row=start_row_offset - 1, end_column=12)

    # Make the text bold
    cell.font = Font(bold=True, color="FF0000")
    cell.alignment = Alignment(horizontal="center", vertical="center")

    df = pd.DataFrame({
        "Department": [""] * 5,
        "Color": [""] * 5,
        "Fund": [""] * 5,
        "Account": [""] * 5,
        "Sub-Account": [""] * 5,
        "Code": [""] * 5,
        "Description": [""] * 5,
    })
    
    write_tables_to_excel(
        wb=wb,
        sheet_name="USER INPUT",
        tables_list=[df],
        start_row=start_row_offset,
        base_table_name="funding_strings",
        buffer=3
    )

    df_1 = pd.DataFrame({
        "EOY Balance": [""] * 5,
        "Income": [""] * 5,
        "Expenditures": [""] * 5,
        "Balance": [""] * 5
    })

    write_tables_to_excel(
        wb=wb,
        sheet_name="USER INPUT",
        tables_list=[df_1],
        start_row=start_row_offset,
        base_table_name="fund_acc",
        start_col=9,
        buffer=3
    )

    # Return the row index after the table and buffer
    return start_row_offset + max(len(df), len(df_1)) + 3  # 3 is the buffer




# Global dataframe to collect all employee names with their categories
all_employee_info = pd.DataFrame(columns=["Name", "Category"])

def process_dfs_based_on_identifier(dfs, desired_columns_map):
    """
    Processes a list of dataframes by identifying their type based on column headers,
    standardizing columns, and optionally modifying content (e.g., adding blank rows).

    Additionally, logs each employee's name along with their matched category in the
    global DataFrame `all_employee_info`.

    Args:
        dfs (list): List of pandas DataFrames representing different employee tables.

    Returns:
        list: List of processed pandas DataFrames.
    """
    global all_employee_info  # allow writing to global DataFrame

    processed_dfs = []

    for df in dfs:
        identifier = df.columns[0]
        raw_identifier = str(identifier).lower()

        matched_key = next(
            (key for key in desired_columns_map if key in raw_identifier),
            None
        )

        if matched_key:
            desired_columns = desired_columns_map[matched_key]

            if "Name" in df.columns:
                df = df.dropna(subset=["Name"])

                #Append names + category to global all_employee_info
                temp_info = df[["Name"]].copy()
                temp_info["Category"] = matched_key
                all_employee_info = pd.concat([all_employee_info, temp_info], ignore_index=True)

            for col in desired_columns:
                if col not in df.columns:
                    df.loc[:, col] = np.nan

            processed_columns = [identifier] + [col for col in desired_columns if col in df.columns]
            df = df[processed_columns]

            if matched_key == "salaried employees":
                df = add_blank_rows(df, 3)

            processed_dfs.append(df)
        else:
            print(f"Warning: No match found for identifier '{raw_identifier}'.")

    return processed_dfs



def add_blank_rows(df, blank_row_count=3):
    """
    Add a specified number of blank rows after each entry in the 'Name' column.
    
    Parameters:
    - df: DataFrame to modify.
    - blank_row_count: Number of blank rows to add after each row. Default is 3.
    
    Returns:
    - Modified DataFrame with blank rows added.
    """
    if "Name" not in df.columns:
        return df  # Skip if the 'Name' column doesn't exist

    new_rows = []
    for _, row in df.iterrows():
        new_rows.append(row)  # Add the original row
        for _ in range(blank_row_count):
            # Add blank rows
            new_rows.append(pd.Series([np.nan] * len(df.columns), index=df.columns))


    return pd.DataFrame(new_rows, columns=df.columns)



def write_tables_to_excel(wb, sheet_name, tables_list, start_row, base_table_name, start_col= 1, buffer=3):
    """
    Write a list of DataFrames to an Excel sheet as tables.

    Args:
        wb (Workbook): Excel workbook object.
        sheet_name (str): Name of the sheet to write to.
        tables_list (list): List of DataFrames to write.
        start_row (int): Starting row for the first table.
        buffer (int): Number of blank rows between tables.
    """
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(title=sheet_name)

    ws = wb[sheet_name]
    current_row = start_row

    for idx, df in enumerate(tables_list):
        # Ensure DataFrame is not empty
        if df.empty:
            print(f"Skipping empty DataFrame at index {idx}")
            continue

        # Ensure DataFrame has unique column headers
        if len(df.columns) != len(set(df.columns)):
            print(f"DataFrame at index {idx} has duplicate column headers. Skipping.")
            continue

        # Use the first column header as the base for the table name
        first_header = str(df.columns[0])
        # Sanitize the table name
        sanitized_header = ''.join(e if e.isalnum() else '_' for e in first_header)
        table_name = f"{base_table_name}Table_{sanitized_header}_{idx + 1}"

        # Define the range for the table
        start_col = start_col  
        end_col = start_col + len(df.columns) - 1
        end_row = current_row + len(df)

        table_range = f"{get_column_letter(start_col)}{current_row}:{get_column_letter(end_col)}{end_row}"

        # Write the DataFrame to the sheet
        for row in dataframe_to_rows(df, index=False, header=True):
            for col_idx, value in enumerate(row, start=start_col):
                ws.cell(row=current_row, column=col_idx, value=value)
            current_row += 1

        # Add buffer
        current_row += buffer

        # Validate the table range before creating the table
        if len(df) > 0 and len(df.columns) > 0:
            try:
                # Create an Excel table
                table = Table(displayName=table_name, ref=table_range)
                style = TableStyleInfo(
                    name="TableStyleMedium9",  # Choose any predefined Excel style
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False,
                )
                table.tableStyleInfo = style
                ws.add_table(table)
                print(f"Successfully added table: {table_name} with range: {table_range}")
            except ValueError as e:
                print(f"Error creating table {table_name}: {e}")
        else:
            print(f"Skipping DataFrame at index {idx} due to invalid dimensions.")

    print(f"All tables written to sheet '{sheet_name}' successfully.")



# This file contains Utility Functions that are used all over the app

import os
import sys
from datetime import datetime
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
import pandas as pd

COLOR_PALETTE = {
    "maroon": "C00000",
    "yellow":"FFC000",

    #Banner Colors
    "lavender":"D5B8EA",
    "pink": "FF8FFF",
    "blue": "BDD7F1",
    "peach": "F9D2BD",
    "green": "D6EDBD",
    "cyan": "D9FFFF",
    
    }

def resource_path(relative_path):
    """ Get the absolute path to a resource, works for dev and for PyInstaller """
    base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base_path, relative_path)


def get_save_name(spreadsheet_name):
    """ This Function forms the new file name using the nomencalture which is {filename_currentDate_cuurentTime)

    Args:
        spreadsheet_name (String): Spreadsheet Base name

    Returns:
        String: Formatted name of file path to be saved
    """
    current_datetime = datetime.now().strftime("%b-%d-%Y_%H-%M-%S")
    return f"{spreadsheet_name}_{current_datetime}.xlsx"


def apply_color(cell, color_name):
    """Applies a predefined color to a cell."""
    color_code = COLOR_PALETTE.get(color_name, "FFFFFF")  # Default to white if not found
    fill = PatternFill(start_color=color_code, end_color=color_code, fill_type="solid")
    cell.fill = fill

def read_excel_table(excel_file, sheet_name, table_range):
    """ This function is used to read the values inside any specified range

    Args:
        excel_file (_type_): Excel spreadsheet path
        sheet_name (_type_): Tab name
        table_range (_type_): Range of the table(can be found using extract ramge function)

    Returns:
        _type_: _description_
    """
    # Load Excel workbook
    wb = load_workbook(excel_file, read_only=True)
    try:
        # Select the worksheet
        ws = wb[sheet_name]

        # Read data from the specified table range into a list of lists
        table_data = []
        for row in ws[table_range]:
            table_data.append([cell.value for cell in row])

        # Convert the list of lists to a DataFrame
        df = pd.DataFrame(table_data[1:], columns=table_data[0])
    finally:
        # Ensure the workbook is closed
        wb.close()

    return df


def extract_table_boundaries(file_path):
    """ Returns table boundaries from a given sheet

    Args:
        file_path (_type_):Path of the Spreadsheet

    Returns:
        _type_: A List of Table Boundaries
    """
    # Load the workbook
    wb = load_workbook(file_path, data_only=True)

    # List of Table Boundaries
    table_ranges= []

    # Loop through all worksheets
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]

        # Check for tables in the sheet
        if hasattr(sheet, 'tables') and sheet.tables:
            for table_name in sheet.tables.keys():
                table = sheet.tables[table_name]  # Retrieve the table object
                if hasattr(table, 'ref'):  # Safely access ref attribute
                    table_range = table.ref
                    table_ranges.append(table_range)

    return table_ranges